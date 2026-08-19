# -*- coding: utf-8 -*-
"""
core/database.py — Capa de almacenamiento SQLite.

Diseño flexible: la estructura de la tabla de mediciones se deriva
automáticamente del parser que se le pase. Un StandardParser genera
columnas (rpm, temp, flujo, rele); un MultiTempParser genera
(rpm, t_tapa_sup, ..., t_amb_inf). El resto del sistema no cambia.

Para una nueva máquina con variables distintas: solo necesitás un
nuevo parser con su display_config y la BD se crea sola.

Tablas que siempre existen (independiente del parser):
    mediciones   — datos en tiempo real + metadatos de ensayo
    alarmas      — registro histórico de eventos de alarma

Thread safety: todas las escrituras pasan por un Lock. Las lecturas
usan conexiones separadas con check_same_thread=False (SQLite WAL).
"""

from __future__ import annotations

import json
import os
import sqlite3
import threading
from datetime import datetime, timedelta, time as dt_time
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.parsers import DataParser


# ===========================================================================
#  CLASE PRINCIPAL
# ===========================================================================
class MedicionesDB:
    """
    Base de datos SQLite para una máquina.

    Uso básico:
        db = MedicionesDB("mi_maquina.db", parser)
        db.save_medicion(data_dict, ensayo="Test A", prueba="Run 1", comentario="")
        dias = db.get_days()
        filas = db.get_mediciones_del_dia("2025-01-15")
        db.export_to_excel("salida.xlsx", dias_seleccionados)
        db.close()
    """

    def __init__(self, db_path: str, parser: DataParser) -> None:
        self._db_path = db_path
        self._parser = parser
        self._lock = threading.Lock()

        # Derivamos el esquema del parser de una vez para siempre
        self._data_columns: list[tuple[str, str, str]] = self._build_columns(parser)

        self._conn: Optional[sqlite3.Connection] = None
        self._connect()
        self._create_tables()

    # -----------------------------------------------------------------------
    # Construcción del esquema dinámico
    # -----------------------------------------------------------------------
    @staticmethod
    def _build_columns(parser: DataParser) -> list[tuple[str, str, str]]:
        """
        Retorna lista de (nombre_columna, tipo_sql, label_display).
        Se deriva del display_config del parser.
        Los campos booleanos (flujo, rele) se almacenan como INTEGER.
        """
        cols = []
        for vd in parser.display_config:
            sql_type = "INTEGER" if vd.is_boolean else "REAL"
            cols.append((vd.key, sql_type, vd.label))
        return cols

    def _columns_ddl(self) -> str:
        """Genera el fragmento DDL de las columnas de datos para CREATE TABLE."""
        return ",\n    ".join(
            f"{name} {sql_type}" for name, sql_type, _ in self._data_columns
        )

    # -----------------------------------------------------------------------
    # Conexión y creación de tablas
    # -----------------------------------------------------------------------
    def _connect(self) -> None:
        os.makedirs(os.path.dirname(os.path.abspath(self._db_path)), exist_ok=True)
        self._conn = sqlite3.connect(
            self._db_path,
            check_same_thread=False,
            detect_types=sqlite3.PARSE_DECLTYPES,
        )
        # WAL mode: permite lecturas concurrentes mientras se escribe
        self._conn.execute("PRAGMA journal_mode=WAL")

    def _create_tables(self) -> None:
        """Crea las tablas si no existen. Migración segura para BDs viejas."""
        with self._lock:
            c = self._conn.cursor()

            # Tabla de mediciones (columnas dinámicas según el parser)
            c.execute(f"""
                CREATE TABLE IF NOT EXISTS mediciones (
                    id        INTEGER PRIMARY KEY AUTOINCREMENT,
                    timestamp TEXT    NOT NULL,
                    fecha     TEXT    NOT NULL,
                    hora      TEXT    NOT NULL,
                    ensayo    TEXT,
                    prueba    TEXT,
                    comentario TEXT,
                    {self._columns_ddl()}
                )
            """)

            # Índices para consultas rápidas por fecha y ensayo
            c.execute("""
                CREATE INDEX IF NOT EXISTS idx_fecha
                ON mediciones (fecha)
            """)
            c.execute("""
                CREATE INDEX IF NOT EXISTS idx_ensayo
                ON mediciones (ensayo, prueba)
            """)

            # Tabla de alarmas (fija, igual para todas las máquinas)
            c.execute("""
                CREATE TABLE IF NOT EXISTS alarmas (
                    id        INTEGER PRIMARY KEY AUTOINCREMENT,
                    timestamp TEXT    NOT NULL,
                    fecha     TEXT    NOT NULL,
                    hora      TEXT    NOT NULL,
                    evento    TEXT,
                    detalle   TEXT,
                    operador  TEXT,
                    ensayo    TEXT,
                    prueba    TEXT
                )
            """)

            # Migración: agregar columnas que puedan faltar en BDs antiguas
            existing = {row[1] for row in c.execute("PRAGMA table_info(mediciones)")}
            for col_name, col_type, _ in self._data_columns:
                if col_name not in existing:
                    try:
                        c.execute(f"ALTER TABLE mediciones ADD COLUMN {col_name} {col_type}")
                    except sqlite3.OperationalError:
                        pass

            self._conn.commit()

    # -----------------------------------------------------------------------
    # Escritura de datos
    # -----------------------------------------------------------------------
    def save_medicion(
        self,
        data: dict,
        ensayo: str = "",
        prueba: str = "",
        comentario: str = "",
        timestamp: Optional[datetime] = None,
    ) -> None:
        """
        Guarda un punto de medición.
        data: el dict que retorna el parser (las claves deben coincidir
              con las columnas del esquema; las que falten quedan NULL).
        """
        if timestamp is None:
            timestamp = datetime.now()

        col_names = [col for col, _, _ in self._data_columns]
        values = [data.get(col) for col in col_names]

        placeholders = ", ".join(["?"] * len(col_names))
        columns_str = ", ".join(col_names)

        sql = f"""
            INSERT INTO mediciones
                (timestamp, fecha, hora, ensayo, prueba, comentario, {columns_str})
            VALUES
                (?, ?, ?, ?, ?, ?, {placeholders})
        """

        row = (
            timestamp.isoformat(),
            timestamp.strftime("%Y-%m-%d"),
            timestamp.strftime("%H:%M:%S"),
            ensayo,
            prueba,
            comentario,
            *values,
        )

        with self._lock:
            self._conn.execute(sql, row)
            self._conn.commit()

    def log_alarma(
        self,
        evento: str,
        detalle: str,
        operador: str = "SISTEMA",
        ensayo: str = "",
        prueba: str = "",
        timestamp: Optional[datetime] = None,
    ) -> None:
        """Registra un evento de alarma en el historial."""
        if timestamp is None:
            timestamp = datetime.now()

        sql = """
            INSERT INTO alarmas
                (timestamp, fecha, hora, evento, detalle, operador, ensayo, prueba)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """
        row = (
            timestamp.isoformat(),
            timestamp.strftime("%Y-%m-%d"),
            timestamp.strftime("%H:%M:%S"),
            evento,
            detalle,
            operador,
            ensayo,
            prueba,
        )
        with self._lock:
            self._conn.execute(sql, row)
            self._conn.commit()

    # -----------------------------------------------------------------------
    # Lectura de datos
    # -----------------------------------------------------------------------
    def get_days(self) -> list[str]:
        """Retorna lista de fechas (YYYY-MM-DD) que tienen mediciones."""
        c = self._conn.cursor()
        c.execute("SELECT DISTINCT fecha FROM mediciones ORDER BY fecha")
        return [row[0] for row in c.fetchall()]

    def get_mediciones_del_dia(
        self,
        fecha: str,
        h_inicio: Optional[dt_time] = None,
        h_fin: Optional[dt_time] = None,
        intervalo_seg: int = 1,
    ) -> list[dict]:
        """
        Retorna mediciones de un día como lista de dicts.
        Soporta filtrado horario e interpolación por intervalo.

        fecha: 'YYYY-MM-DD'
        intervalo_seg: 1 = todos los puntos; N = promedia en ventanas de N segundos
        """
        c = self._conn.cursor()
        col_names = [col for col, _, _ in self._data_columns]
        cols_sql = ", ".join(col_names)

        query = f"""
            SELECT timestamp, ensayo, prueba, comentario, {cols_sql}
            FROM mediciones
            WHERE fecha = ?
        """
        params: list = [fecha]

        if h_inicio:
            query += " AND time(timestamp) >= ?"
            params.append(h_inicio.strftime("%H:%M:%S"))
        if h_fin:
            query += " AND time(timestamp) <= ?"
            params.append(h_fin.strftime("%H:%M:%S"))

        query += " ORDER BY timestamp ASC"
        c.execute(query, params)
        rows = c.fetchall()

        if not rows:
            return []

        # Convertir a lista de dicts
        result = []
        for row in rows:
            ts_str, ensayo, prueba, comentario, *data_vals = row
            entry = {
                "timestamp": datetime.fromisoformat(ts_str),
                "ensayo": ensayo,
                "prueba": prueba,
                "comentario": comentario,
            }
            for col_name, val in zip(col_names, data_vals):
                entry[col_name] = val
            result.append(entry)

        if intervalo_seg <= 1:
            return result

        # Reducción por ventana de tiempo
        return self._reduce_by_interval(result, intervalo_seg)

    @staticmethod
    def _reduce_by_interval(rows: list[dict], intervalo_seg: int) -> list[dict]:
        """Agrupa filas en ventanas de tiempo y promedia los valores numéricos."""
        if not rows:
            return []

        groups: list[list[dict]] = []
        current_group: list[dict] = [rows[0]]
        group_start: datetime = rows[0]["timestamp"]

        for row in rows[1:]:
            if (row["timestamp"] - group_start).total_seconds() < intervalo_seg:
                current_group.append(row)
            else:
                groups.append(current_group)
                current_group = [row]
                group_start = row["timestamp"]
        groups.append(current_group)

        result = []
        for group in groups:
            representative = dict(group[0])  # metadatos del primer punto
            representative["timestamp"] = group[len(group) // 2]["timestamp"]

            # Promediar columnas numéricas
            for key in group[0]:
                if isinstance(group[0][key], (int, float)) and group[0][key] is not None:
                    vals = [r[key] for r in group if r[key] is not None]
                    if vals:
                        representative[key] = sum(vals) / len(vals)
            result.append(representative)

        return result

    def get_alarmas(self) -> list[dict]:
        """Retorna todo el registro de alarmas, del más reciente al más viejo."""
        c = self._conn.cursor()
        c.execute("""
            SELECT timestamp, fecha, hora, evento, detalle, operador, ensayo, prueba
            FROM alarmas
            ORDER BY id DESC
        """)
        cols = ["timestamp", "fecha", "hora", "evento", "detalle", "operador", "ensayo", "prueba"]
        return [dict(zip(cols, row)) for row in c.fetchall()]

    def get_runs(self) -> list[dict]:
        """Retorna combinaciones únicas de (ensayo, prueba, fecha) para el visor."""
        c = self._conn.cursor()
        c.execute("""
            SELECT DISTINCT ensayo, prueba, fecha
            FROM mediciones
            ORDER BY fecha DESC, ensayo, prueba
        """)
        return [{"ensayo": r[0], "prueba": r[1], "fecha": r[2]} for r in c.fetchall()]

    # -----------------------------------------------------------------------
    # Exportación a Excel
    # -----------------------------------------------------------------------
    def export_to_excel(
        self,
        filename: str,
        fechas: list[str],
        ensayos_meta: dict,        # {nombre_ensayo: [{atributo, valor}, ...]}
        h_inicio: Optional[dt_time] = None,
        h_fin: Optional[dt_time] = None,
        intervalo_seg: int = 1,
    ) -> str:
        """
        Exporta datos a Excel en formato HORIZONTAL (Bloques por Prueba).
        """
        wb = Workbook()
        ws_data = wb.active
        ws_data.title = "Datos de Medición"

        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        col_labels = [lbl for _, _, lbl in self._data_columns]
        col_names = [name for name, _, _ in self._data_columns]

        col_offset = 1

        for fecha in fechas:
            filas = self.get_mediciones_del_dia(fecha, h_inicio, h_fin, intervalo_seg)
            if not filas:
                continue

            # 1. Agrupar filas por (Ensayo, Prueba)
            ensayos_en_dia = {}
            for fila in filas:
                ens = fila.get("ensayo") or "Sin Ensayo"
                pr = fila.get("prueba") or "Sin Prueba"
                clave = f"E:{ens}|P:{pr}"
                if clave not in ensayos_en_dia:
                    ensayos_en_dia[clave] = {"ensayo": ens, "prueba": pr, "filas": []}
                ensayos_en_dia[clave]["filas"].append(fila)

            # 2. Calcular la fila donde empiezan los datos (depende del ensayo con más atributos)
            max_attrs = 0
            for info in ensayos_en_dia.values():
                ens = info["ensayo"]
                if ens in ensayos_meta:
                    max_attrs = max(max_attrs, len(ensayos_meta[ens]))

            # Fila base para los encabezados de datos (baja según los atributos)
            data_header_row = 1 + 3 + (max_attrs + 1 if max_attrs > 0 else 0) + 1

            # 3. Escribir cada grupo (Ensayo/Prueba) en columnas separadas (lado a lado)
            for info in ensayos_en_dia.values():
                current_row = 1
                header_text = f"Fecha: {fecha}\nEnsayo: {info['ensayo']}\nPrueba: {info['prueba']}"

                # Ancho del bloque: 1 (Hora) + Variables + 1 (Comentario)
                block_width = 1 + len(col_names) + 1

                # Título principal del bloque
                ws_data.merge_cells(start_row=current_row, start_column=col_offset,
                                    end_row=current_row+2, end_column=col_offset + block_width - 1)
                cell = ws_data.cell(row=current_row, column=col_offset, value=header_text)
                cell.font = bold_font
                cell.alignment = center_align
                
                # Le damos un colorcito de fondo suave para separar visualmente
                cell.fill = PatternFill("solid", fgColor="E8F4F8") 
                current_row += 3

                # Escribir Atributos debajo del título general
                if info['ensayo'] != "Sin Ensayo" and info['ensayo'] in ensayos_meta:
                    ws_data.cell(row=current_row, column=col_offset, value="Atributo").font = bold_font
                    ws_data.cell(row=current_row, column=col_offset+1, value="Valor").font = bold_font
                    current_row += 1
                    for attr in ensayos_meta[info['ensayo']]:
                        ws_data.cell(row=current_row, column=col_offset, value=attr.get("atributo", ""))
                        ws_data.cell(row=current_row, column=col_offset+1, value=attr.get("valor", ""))
                        current_row += 1

                # Escribir Encabezados de Datos (dinámico según la máquina)
                data_headers = ["Hora"] + col_labels + ["Comentario"]
                for i, header in enumerate(data_headers):
                    c = ws_data.cell(row=data_header_row, column=col_offset + i, value=header)
                    c.font = bold_font
                    c.alignment = center_align
                    # Ensanchar columnas
                    ws_data.column_dimensions[get_column_letter(col_offset + i)].width = 15

                # Escribir Filas de Datos
                data_current_row = data_header_row + 1
                for fila in info["filas"]:
                    # Hora
                    ws_data.cell(row=data_current_row, column=col_offset).value = fila["timestamp"].strftime("%H:%M:%S")

                    # Variables dinámicas de los sensores
                    for i, col in enumerate(col_names):
                        ws_data.cell(row=data_current_row, column=col_offset + 1 + i).value = fila.get(col)

                    # Comentario
                    com = fila.get("comentario", "")
                    ws_data.cell(row=data_current_row, column=col_offset + 1 + len(col_names)).value = com if com else ""

                    data_current_row += 1

                # Separación entre bloques: sumar el ancho del bloque + 1 columna en blanco
                col_offset += block_width + 1

        # Congelar paneles para poder scrollear sin perder los encabezados
        ws_data.freeze_panes = ws_data.cell(row=data_header_row + 1, column=1)

        # Hoja de alarmas
        alarmas = self.get_alarmas()
        if alarmas:
            ws_alarm = wb.create_sheet(title="Registro de Alarmas")
            alarm_headers = ["Fecha", "Hora", "Evento", "Detalle", "Operador", "Ensayo", "Prueba"]
            for i, h in enumerate(alarm_headers, 1):
                c = ws_alarm.cell(row=1, column=i, value=h)
                c.font = bold_font
                c.alignment = center_align
                ws_alarm.column_dimensions[get_column_letter(i)].width = 20

            for row_i, a in enumerate(alarmas, 2):
                row_data = [a["fecha"], a["hora"], a["evento"], a["detalle"], a["operador"], a["ensayo"], a["prueba"]]
                for col_i, val in enumerate(row_data, 1):
                    ws_alarm.cell(row=row_i, column=col_i, value=val)

        wb.save(filename)
        return filename

    # -----------------------------------------------------------------------
    # Mantenimiento
    # -----------------------------------------------------------------------
    def change_db(self, new_path: str) -> None:
        """Cambia la base de datos en caliente. Para uso desde la UI."""
        self.close()
        self._db_path = new_path
        self._connect()
        self._create_tables()

    def close(self) -> None:
        """Cierra la conexión limpiamente."""
        with self._lock:
            if self._conn:
                try:
                    self._conn.close()
                except Exception:
                    pass
                self._conn = None

    @property
    def db_path(self) -> str:
        return self._db_path

    @property
    def column_names(self) -> list[str]:
        """Nombres de las columnas de datos (para la UI)."""
        return [name for name, _, _ in self._data_columns]
